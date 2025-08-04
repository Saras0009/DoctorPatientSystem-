import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Excel file setup
file_path = "user_data.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    sheet = wb.active
    sheet.append([
        "User Type", "First Name", "Last Name", "Profile Pic", "Username", "Email",
        "Password", "Address Line 1", "City", "State", "Pincode"
    ])
    wb.save(file_path)

# Clear all signup fields
def clear_signup():
    for entry in [fname_entry, lname_entry, profile_entry, username_entry, email_entry,
                  pwd_entry, cpwd_entry, address1_entry, city_entry, state_entry, pincode_entry]:
        entry.delete(0, tk.END)

# Signup function
def signup():
    user_type = user_type_var.get()
    fname = fname_entry.get()
    lname = lname_entry.get()
    profile = profile_entry.get()
    username = username_entry.get()
    email = email_entry.get()
    pwd = pwd_entry.get()
    cpwd = cpwd_entry.get()
    address1 = address1_entry.get()
    city = city_entry.get()
    state = state_entry.get()
    pincode = pincode_entry.get()

    if pwd != cpwd:
        messagebox.showerror("Error", "Password and Confirm Password must match.")
        return

    wb = load_workbook(file_path)
    sheet = wb.active
    sheet.append([
        user_type, fname, lname, profile, username, email, pwd,
        address1, city, state, pincode
    ])
    wb.save(file_path)
    messagebox.showinfo("Success", f"{user_type} signed up successfully!")
    clear_signup()

# Login function
def login():
    username = login_username.get()
    pwd = login_pwd.get()

    wb = load_workbook(file_path)
    sheet = wb.active

    for i in range(2, sheet.max_row + 1):
        stored_username = sheet.cell(i, 5).value
        stored_pwd = sheet.cell(i, 7).value

        if username == stored_username and pwd == stored_pwd:
            user_type = sheet.cell(i, 1).value
            fname = sheet.cell(i, 2).value
            lname = sheet.cell(i, 3).value
            profile = sheet.cell(i, 4).value
            email = sheet.cell(i, 6).value
            address1 = sheet.cell(i, 8).value
            city = sheet.cell(i, 9).value
            state = sheet.cell(i, 10).value
            pincode = sheet.cell(i, 11).value

            msg = f"""Welcome, {user_type}!

Full Name: {fname} {lname}
Username: {username}
Email: {email}
Address: {address1}, {city}, {state} - {pincode}
Profile Pic Path: {profile}
"""
            messagebox.showinfo(f"{user_type} Dashboard", msg)
            return

    messagebox.showerror("Error", "Invalid username or password.")

# --------------------------- GUI Setup ---------------------------

root = tk.Tk()
root.title("Signup/Login System")
root.geometry("600x750")
root.configure(bg="white")

# Title
tk.Label(root, text="Signup Form", font=("Arial", 16, "bold"), bg="orange", fg="white").grid(row=0, column=0, columnspan=2, pady=10)

# User Type
user_type_var = tk.StringVar()
user_type_var.set("Patient")
tk.Label(root, text="User Type", bg="white").grid(row=1, column=0, pady=5)
tk.OptionMenu(root, user_type_var, "Patient", "Doctor").grid(row=1, column=1, pady=5)

# Signup Fields
tk.Label(root, text="First Name", bg="white").grid(row=2, column=0, pady=5)
fname_entry = tk.Entry(root); fname_entry.grid(row=2, column=1, pady=5)

tk.Label(root, text="Last Name", bg="white").grid(row=3, column=0, pady=5)
lname_entry = tk.Entry(root); lname_entry.grid(row=3, column=1, pady=5)

tk.Label(root, text="Profile Pic (path)", bg="white").grid(row=4, column=0, pady=5)
profile_entry = tk.Entry(root); profile_entry.grid(row=4, column=1, pady=5)

tk.Label(root, text="Username", bg="white").grid(row=5, column=0, pady=5)
username_entry = tk.Entry(root); username_entry.grid(row=5, column=1, pady=5)

tk.Label(root, text="Email", bg="white").grid(row=6, column=0, pady=5)
email_entry = tk.Entry(root); email_entry.grid(row=6, column=1, pady=5)

tk.Label(root, text="Password", bg="white").grid(row=7, column=0, pady=5)
pwd_entry = tk.Entry(root, show="*"); pwd_entry.grid(row=7, column=1, pady=5)

tk.Label(root, text="Confirm Password", bg="white").grid(row=8, column=0, pady=5)
cpwd_entry = tk.Entry(root, show="*"); cpwd_entry.grid(row=8, column=1, pady=5)

tk.Label(root, text="Address Line 1", bg="white").grid(row=9, column=0, pady=5)
address1_entry = tk.Entry(root); address1_entry.grid(row=9, column=1, pady=5)

tk.Label(root, text="City", bg="white").grid(row=10, column=0, pady=5)
city_entry = tk.Entry(root); city_entry.grid(row=10, column=1, pady=5)

tk.Label(root, text="State", bg="white").grid(row=11, column=0, pady=5)
state_entry = tk.Entry(root); state_entry.grid(row=11, column=1, pady=5)

tk.Label(root, text="Pincode", bg="white").grid(row=12, column=0, pady=5)
pincode_entry = tk.Entry(root); pincode_entry.grid(row=12, column=1, pady=5)

# Signup Button
tk.Button(root, text="Signup", bg="black", fg="white", command=signup).grid(row=13, column=0, columnspan=2, pady=10)

# Login Section Title
tk.Label(root, text="Login Section", font=("Arial", 16, "bold"), bg="orange", fg="white").grid(row=14, column=0, columnspan=2, pady=10)

# Login Fields
tk.Label(root, text="Username", bg="white").grid(row=15, column=0, pady=5)
login_username = tk.Entry(root); login_username.grid(row=15, column=1, pady=5)

tk.Label(root, text="Password", bg="white").grid(row=16, column=0, pady=5)
login_pwd = tk.Entry(root, show="*"); login_pwd.grid(row=16, column=1, pady=5)

# Login Button
tk.Button(root, text="Login", bg="black", fg="white", command=login).grid(row=17, column=0, columnspan=2, pady=10)

root.mainloop()
